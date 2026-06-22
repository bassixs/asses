from __future__ import annotations

import asyncio
from collections import defaultdict
from dataclasses import dataclass
import json
import logging
from pathlib import Path
import re
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pydantic import BaseModel, Field, ValidationError, field_validator

from bot.config import settings
from bot.services.llm_json import LLMJSONError, complete_json_openai_compatible
from bot.services.exercise_context import build_exercise_analysis_block
from bot.services.role_labeling import extract_participant_transcript
from bot.services.yandex_gpt import complete_json

logger = logging.getLogger(__name__)

IndicatorStatus = Literal["+", "-", "НЗ"]


class NotebookProcessingError(RuntimeError):
    pass


@dataclass(frozen=True)
class IndicatorRow:
    indicator_id: str
    row: int
    competence: str
    indicator: str
    status_col: int = 4
    comment_col: int = 5
    level_col: int = 6
    # Row of the behavioural subgroup header (its yellow level cell), if the notebook
    # groups indicators into subgroups within a competence.
    group_row: int | None = None


class IndicatorEvidence(BaseModel):
    timestamp: str | None = None
    quote: str


class IndicatorAnalysis(BaseModel):
    indicator_id: str
    status: IndicatorStatus
    evidence: list[IndicatorEvidence] = Field(default_factory=list)
    comment: str = ""
    observable_reason: str = ""


class NotebookAnalysisReport(BaseModel):
    role_summary: str = ""
    participant_summary: str = ""
    results: list[IndicatorAnalysis]

    @field_validator("results")
    @classmethod
    def results_must_not_be_empty(cls, value: list[IndicatorAnalysis]) -> list[IndicatorAnalysis]:
        if not value:
            raise ValueError("results must not be empty")
        return value


def extract_notebook_indicators(workbook_path: Path) -> list[IndicatorRow]:
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    indicators: list[IndicatorRow] = []
    current_competence = ""
    current_group_row: int | None = None
    layout = _detect_notebook_layout(sheet)

    for row in range(1, sheet.max_row + 1):
        competence_raw = _cell_text(sheet.cell(row=row, column=layout["competence_col"]).value)
        indicator = _cell_text(sheet.cell(row=row, column=layout["indicator_col"]).value)
        status_header = _cell_text(sheet.cell(row=row, column=layout["status_col"]).value)

        if competence_raw and not _looks_like_column_header(competence_raw) and not _looks_like_level(competence_raw):
            current_competence = competence_raw
            current_group_row = None

        if not indicator or _looks_like_column_header(indicator) or _looks_like_column_header(status_header):
            # A subgroup header repeats the "Проявления"/"Комментарии" column titles and
            # carries the subgroup name in the indicator column — its row holds the
            # subgroup's yellow level cell.
            if indicator and not _looks_like_column_header(indicator) and _looks_like_column_header(status_header):
                current_group_row = row
            continue

        competence = "" if _looks_like_level(competence_raw) else competence_raw
        competence = competence or current_competence
        if not competence:
            competence = "Компетенция не указана"

        indicators.append(
            IndicatorRow(
                indicator_id=f"I{len(indicators) + 1:03d}",
                row=row,
                competence=competence,
                indicator=indicator,
                status_col=layout["status_col"],
                comment_col=layout["comment_col"],
                level_col=layout["level_col"],
                group_row=current_group_row,
            )
        )

    if not indicators:
        raise NotebookProcessingError(
            "В блокноте не найдены поведенческие индикаторы. "
            "Проверьте, что рядом с ними есть колонки 'Проявления' и 'Комментарии'."
        )

    return indicators


async def analyze_notebook_indicators(
    *,
    transcript: str,
    indicators: list[IndicatorRow],
    exercise_name: str | None = None,
    exercise_instructions: str | None = None,
) -> NotebookAnalysisReport:
    system_prompt = _build_notebook_system_prompt()
    participant_transcript = extract_participant_transcript(transcript)
    exercise_context = build_exercise_analysis_block(exercise_name, exercise_instructions)
    if exercise_context:
        source = "uploaded instructions" if exercise_instructions else "bundled library"
        logger.info("Notebook analysis using exercise context for '%s' (%s)", exercise_name, source)

    batch_size = max(1, settings.notebook_analysis_batch_size)
    batches = [indicators[i : i + batch_size] for i in range(0, len(indicators), batch_size)]

    semaphore = asyncio.Semaphore(max(1, settings.analysis_llm_max_concurrency))

    async def run_batch(batch_index: int, batch: list[IndicatorRow]) -> NotebookAnalysisReport | None:
        async with semaphore:
            try:
                return await _analyze_indicator_batch(
                    system_prompt=system_prompt,
                    transcript=participant_transcript,
                    indicators=batch,
                    batch_index=batch_index,
                    batch_count=len(batches),
                    exercise_context=exercise_context,
                )
            except (LLMJSONError, NotebookProcessingError) as exc:
                # One bad batch should not fail the whole exercise; its indicators are
                # marked "НЗ" later by _ensure_all_indicators.
                logger.error("Notebook analysis batch %s/%s failed: %s", batch_index, len(batches), exc)
                return None

    reports = await asyncio.gather(
        *(run_batch(index, batch) for index, batch in enumerate(batches, start=1))
    )

    all_results: list[IndicatorAnalysis] = []
    role_summary = ""
    participant_summary = ""
    succeeded = 0

    for report in reports:
        if report is None:
            continue
        succeeded += 1
        all_results.extend(report.results)
        if not role_summary and report.role_summary:
            role_summary = report.role_summary
        if not participant_summary and report.participant_summary:
            participant_summary = report.participant_summary

    if succeeded == 0:
        raise NotebookProcessingError(
            f"Не удалось получить оценку блокнота ни по одному батчу ({len(batches)} шт.). "
            "Подробности ошибок — в логе сервера.",
        )

    merged = NotebookAnalysisReport(
        role_summary=role_summary,
        participant_summary=participant_summary,
        results=all_results,
    )
    return _ensure_all_indicators(merged, indicators)


async def _analyze_indicator_batch(
    *,
    system_prompt: str,
    transcript: str,
    indicators: list[IndicatorRow],
    batch_index: int,
    batch_count: int,
    exercise_context: str = "",
) -> NotebookAnalysisReport:
    payload = [
        {
            "indicator_id": item.indicator_id,
            "competence": item.competence,
            "indicator": item.indicator,
        }
        for item in indicators
    ]
    user_prompt = _build_notebook_user_prompt(
        transcript=transcript,
        indicators=payload,
        exercise_context=exercise_context,
    )
    logger.info("Notebook analysis batch %s/%s: indicators=%s", batch_index, batch_count, len(indicators))

    if settings.analysis_llm_provider.lower().strip() == "yandex":
        raw = await complete_json(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            json_schema=NotebookAnalysisReport.model_json_schema(),
            temperature=0.05,
            max_tokens=settings.analysis_llm_max_tokens,
        )
    else:
        raw = await complete_json_openai_compatible(
            provider=settings.analysis_llm_provider,
            model=settings.analysis_llm_model,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            temperature=settings.analysis_llm_temperature,
            max_tokens=settings.analysis_llm_max_tokens,
            timeout_seconds=settings.analysis_llm_timeout_seconds,
            json_mode=settings.analysis_llm_json_mode,
        )

    try:
        return _coerce_notebook_report(raw)
    except (ValidationError, NotebookProcessingError) as exc:
        logger.error(
            "Invalid notebook analysis JSON (batch %s/%s): %s; raw=%s",
            batch_index,
            batch_count,
            exc,
            _truncate_raw(raw),
        )
        raise NotebookProcessingError(
            f"Модель вернула некорректную структуру оценки блокнота (батч {batch_index}/{batch_count}).",
        ) from exc


# Wrapper keys the model has been observed to use instead of the expected "results".
_RESULT_KEYS = ("results", "observations", "indicators", "segments", "items", "data", "evaluations")
# Field names the model uses for the indicator id.
_ID_KEYS = ("indicator_id", "indicatorId", "indicator_code", "id", "code", "number")
_VALID_STATUSES = {"+", "-", "НЗ"}


def _truncate_raw(raw: Any, limit: int = 1200) -> str:
    try:
        text = json.dumps(raw, ensure_ascii=False)
    except (TypeError, ValueError):
        text = str(raw)
    return text[:limit]


def _coerce_notebook_report(raw: Any) -> NotebookAnalysisReport:
    """Normalize the loosely-structured LLM JSON into NotebookAnalysisReport.

    Tolerates: a bare list; a dict wrapped under results/observations/indicators/etc.;
    a dict keyed by indicator id ({"I001": {...}}); evidence as a plain string; null
    comment/observable_reason; and several id field names.
    """
    role_summary = ""
    participant_summary = ""
    if isinstance(raw, dict):
        role_summary = _as_text(raw.get("role_summary"))
        participant_summary = _as_text(raw.get("participant_summary"))

    items = _extract_items(raw)
    results: list[IndicatorAnalysis] = []
    for item in items:
        analysis = _coerce_indicator(item)
        if analysis is not None:
            results.append(analysis)

    if not results:
        raise NotebookProcessingError("Не удалось извлечь ни одного индикатора из ответа модели.")

    return NotebookAnalysisReport(
        role_summary=role_summary,
        participant_summary=participant_summary,
        results=results,
    )


def _extract_items(raw: Any) -> list[dict[str, Any]]:
    if isinstance(raw, list):
        return [item for item in raw if isinstance(item, dict)]
    if not isinstance(raw, dict):
        return []

    for key in _RESULT_KEYS:
        if key in raw:
            items = _as_item_list(raw[key])
            if items:
                return items

    # The dict may itself be a single indicator...
    if _looks_like_indicator(raw):
        return [raw]
    # ...or a map of indicator_id -> analysis at the top level.
    return _as_item_list(raw)


def _as_item_list(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if isinstance(value, dict):
        items: list[dict[str, Any]] = []
        for key, entry in value.items():
            if isinstance(entry, dict):
                item = dict(entry)
                if not _looks_like_indicator(item):
                    item.setdefault("indicator_id", key)
                items.append(item)
        return items
    return []


def _looks_like_indicator(item: dict[str, Any]) -> bool:
    return any(item.get(key) for key in _ID_KEYS)


def _coerce_indicator(item: Any) -> IndicatorAnalysis | None:
    if not isinstance(item, dict):
        return None
    indicator_id = next((item[key] for key in _ID_KEYS if item.get(key)), None)
    if not indicator_id:
        return None
    status = _as_text(item.get("status"))
    if status not in _VALID_STATUSES:
        status = "НЗ"
    evidence = _coerce_evidence(item.get("evidence")) if status == "+" else []
    return IndicatorAnalysis(
        indicator_id=str(indicator_id),
        status=status,  # type: ignore[arg-type]
        evidence=evidence,
        comment=_as_text(item.get("comment")),
        observable_reason=_as_text(item.get("observable_reason")),
    )


_ROLE_PREFIX_RE = re.compile(r"^\s*(?:участник|ведущий|наблюдатель)\s*:\s*", re.IGNORECASE)


def _clean_quote(text: str) -> str:
    """Strip a leading role label ("Участник:"/"Ведущий:") the LLM sometimes glues
    onto the quote. It is not part of the speech, and it breaks timestamp matching."""
    cleaned = _ROLE_PREFIX_RE.sub("", text or "").strip()
    return cleaned.strip("«»\"' ").strip()


def _coerce_evidence(value: Any) -> list[IndicatorEvidence]:
    if not value:
        return []
    if isinstance(value, str):
        text = _clean_quote(value)
        return [IndicatorEvidence(quote=text)] if text else []
    if isinstance(value, dict):
        value = [value]
    if not isinstance(value, list):
        return []
    out: list[IndicatorEvidence] = []
    for entry in value:
        if isinstance(entry, str):
            text = _clean_quote(entry)
            if text:
                out.append(IndicatorEvidence(quote=text))
        elif isinstance(entry, dict):
            quote = _clean_quote(_as_text(entry.get("quote") or entry.get("text")))
            if quote:
                timestamp = _as_text(entry.get("timestamp")) or None
                out.append(IndicatorEvidence(quote=quote, timestamp=timestamp))
    return out


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def fill_observer_notebook(
    *,
    input_path: Path,
    output_path: Path,
    indicators: list[IndicatorRow],
    report: NotebookAnalysisReport,
) -> dict[str, object]:
    workbook = load_workbook(input_path)
    sheet = workbook.active
    result_by_id = {item.indicator_id: item for item in report.results}
    indicator_by_row = {item.row: item for item in indicators}

    for indicator in indicators:
        result = result_by_id[indicator.indicator_id]
        sheet.cell(row=indicator.row, column=indicator.status_col).value = result.status
        sheet.cell(row=indicator.row, column=indicator.comment_col).value = _format_comment(result)
        sheet.cell(row=indicator.row, column=indicator.comment_col).alignment = Alignment(wrap_text=True, vertical="top")

    levels = _calculate_competency_levels(indicators, result_by_id)
    _write_competency_levels(sheet, indicator_by_row, levels)
    _write_subgroup_levels(sheet, indicators, result_by_id)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "role_summary": report.role_summary,
        "participant_summary": report.participant_summary,
        "indicator_count": len(indicators),
        "levels": levels,
        "indicators": [
            {
                "indicator_id": item.indicator_id,
                "competence": item.competence,
                "indicator": item.indicator,
                "row": item.row,
            }
            for item in indicators
        ],
        "results": [item.model_dump() for item in report.results],
    }


def _build_notebook_system_prompt() -> str:
    return """
Ты — эксперт ассессмент-центра и обученный наблюдатель.
Твоя задача — заполнить блокнот наблюдателя по транскрипту упражнения.

Правила:
1. Оценивай только реплики оцениваемого участника ассессмент-центра.
2. Не используй реплики наблюдателя, ведущего или ролевого игрока как evidence проявления.
3. Во входном транскрипте уже должен быть оставлен оцениваемый участник. Если всё же встретятся роли, используй только строки "Участник:".
4. Для каждого индикатора верни один статус:
   "+" — есть хотя бы одна релевантная цитата участника;
   "-" — индикатор мог наблюдаться в упражнении, но проявления участника нет;
   "НЗ" — индикатор объективно не мог наблюдаться в этом упражнении.
5. Каждый "+" должен иметь точную цитату. Цитата — это ДОСЛОВНЫЕ слова участника из транскрипта (слово в слово). Если в транскрипте есть таймкод, верни его.
   ЗАПРЕЩЕНО брать в качестве цитаты примеры из формулировки самого индикатора (текст в скобках или кавычках внутри описания индикатора) — это НЕ речь участника, а пояснение для наблюдателя. Если дословной цитаты участника в транскрипте нет — ставь "-", а не "+".
6. Для "-" evidence должен быть пустым.
7. Для "НЗ" evidence должен быть пустым, а observable_reason должен объяснять, почему ситуация не позволяла замерить индикатор.
8. Не придумывай цитаты, таймкоды, действия и роли.
9. Если дан "Контекст упражнения", опирайся на него, решая, мог ли индикатор наблюдаться: если упражнение по сценарию не создаёт ситуацию для индикатора — ставь "НЗ", а не "-".
10. Возвращай только валидный JSON по схеме.
""".strip()


def _build_notebook_user_prompt(
    *,
    transcript: str,
    indicators: list[dict[str, str]],
    exercise_context: str = "",
) -> str:
    context_block = f"{exercise_context}\n\n" if exercise_context else ""
    return f"""
{context_block}Транскрипт упражнения:
{transcript}

Поведенческие индикаторы из блокнота наблюдателя:
{indicators}
""".strip()


def _ensure_all_indicators(
    report: NotebookAnalysisReport,
    indicators: list[IndicatorRow],
) -> NotebookAnalysisReport:
    expected = {item.indicator_id for item in indicators}
    existing = {item.indicator_id for item in report.results}
    missing = expected - existing
    if not missing:
        return report

    patched = list(report.results)
    for indicator_id in sorted(missing):
        patched.append(
            IndicatorAnalysis(
                indicator_id=indicator_id,
                status="НЗ",
                evidence=[],
                comment="Индикатор не был возвращен моделью.",
                observable_reason="Автоматически помечено как незамер из-за неполного ответа модели.",
            )
        )
    return NotebookAnalysisReport(
        role_summary=report.role_summary,
        participant_summary=report.participant_summary,
        results=patched,
    )


def _calculate_competency_levels(
    indicators: list[IndicatorRow],
    result_by_id: dict[str, IndicatorAnalysis],
) -> dict[str, dict[str, float | int]]:
    grouped: dict[str, list[IndicatorStatus]] = defaultdict(list)
    for indicator in indicators:
        grouped[indicator.competence].append(result_by_id[indicator.indicator_id].status)

    levels: dict[str, dict[str, float | int]] = {}
    for competence, statuses in grouped.items():
        observed = [status for status in statuses if status != "НЗ"]
        plus_count = sum(1 for status in observed if status == "+")
        observed_count = len(observed)
        percent = (plus_count / observed_count * 100) if observed_count else 0.0
        levels[competence] = {
            "plus_count": plus_count,
            "observed_count": observed_count,
            "percent": round(percent, 2),
            "level": _level_from_percent(percent) if observed_count else 0,
        }
    return levels


def _write_competency_levels(
    sheet,
    indicator_by_row: dict[int, IndicatorRow],
    levels: dict[str, dict[str, float | int]],
) -> None:
    written: set[str] = set()
    for row in range(1, sheet.max_row + 1):
        for column in (1, 2):
            competence = _cell_text(sheet.cell(row=row, column=column).value)
            if competence in levels and competence not in written:
                level_col = _level_column_for_competence(indicator_by_row, competence)
                sheet.cell(row=row, column=level_col).value = levels[competence]["level"]
                written.add(competence)

    for row in range(1, sheet.max_row + 1):
        indicator = indicator_by_row.get(row)
        if indicator and indicator.competence not in written:
            sheet.cell(row=row, column=indicator.level_col).value = levels[indicator.competence]["level"]
            written.add(indicator.competence)


def _write_subgroup_levels(
    sheet,
    indicators: list[IndicatorRow],
    result_by_id: dict[str, IndicatorAnalysis],
) -> None:
    """Write a level into each behavioural-subgroup yellow cell (same formula, per subgroup)."""
    grouped: dict[int, list[IndicatorStatus]] = defaultdict(list)
    level_col_by_group: dict[int, int] = {}
    for indicator in indicators:
        if indicator.group_row is None:
            continue
        grouped[indicator.group_row].append(result_by_id[indicator.indicator_id].status)
        level_col_by_group[indicator.group_row] = indicator.level_col

    for group_row, statuses in grouped.items():
        observed = [status for status in statuses if status != "НЗ"]
        percent = (sum(1 for s in observed if s == "+") / len(observed) * 100) if observed else 0.0
        level = _level_from_percent(percent) if observed else 0
        sheet.cell(row=group_row, column=level_col_by_group[group_row]).value = level


def _level_from_percent(percent: float) -> float:
    if percent >= 90:
        return 3
    if percent >= 80:
        return 2.5
    if percent >= 60:
        return 2
    if percent >= 40:
        return 1.5
    if percent >= 20:
        return 1
    if percent >= 5:
        return 0.5
    return 0


_TS_WORD_RE = re.compile(r"\w+", re.UNICODE)


def verify_evidence_quotes(report: NotebookAnalysisReport, transcript: str) -> None:
    """Drop "+" evidence whose quote is not actually present in the transcript.

    The analysis LLM sometimes lifts the example phrasing from the indicator's own
    description (text in parentheses) and passes it off as a participant quote. Such
    quotes do not appear verbatim in the recording, so we verify each "+" quote against
    the transcript word stream and drop fabricated ones. If a "+" loses all evidence,
    it is downgraded to "-" (no comment), which keeps the level math honest per ТЗ.
    """
    stream = [w.lower() for w in _TS_WORD_RE.findall(transcript or "")]
    if not stream:
        logger.info("Evidence verification skipped: empty transcript")
        return

    dropped = 0
    downgraded = 0
    for result in report.results:
        if result.status != "+":
            continue
        kept = [item for item in result.evidence if _quote_in_transcript(item.quote, stream)]
        dropped += len(result.evidence) - len(kept)
        if kept:
            result.evidence = kept
            continue
        # No verifiable quote left — this "+" was based on a fabricated/paraphrased quote.
        result.status = "-"
        result.evidence = []
        result.comment = ""
        downgraded += 1
    logger.info("Evidence verification: dropped=%s fabricated quotes, downgraded=%s '+'→'-'", dropped, downgraded)


def _quote_in_transcript(quote: str, stream: list[str]) -> bool:
    needle = [w.lower() for w in _TS_WORD_RE.findall(quote)]
    if not needle:
        return False
    run = _longest_word_run(needle, stream)
    if len(needle) <= 3:
        # Very short quote: require the whole thing to appear (min 2 words).
        return run >= max(2, len(needle))
    if len(needle) <= 5:
        return run >= 3
    # Longer quotes: a genuine transcript quote has at least a 4-word verbatim span.
    return run >= 4


def _longest_word_run(needle: list[str], stream: list[str]) -> int:
    best = 0
    for i in range(len(stream)):
        run = 0
        while run < len(needle) and i + run < len(stream) and stream[i + run] == needle[run]:
            run += 1
        if run > best:
            best = run
            if best == len(needle):
                break
    return best


def attach_evidence_timestamps(report: NotebookAnalysisReport, segments: list[dict[str, Any]]) -> None:
    """Fill evidence.timestamp ([мм:сс]) by matching each quote to the STT segments."""
    if not segments:
        logger.info("Evidence timestamps skipped: no transcript segments stored")
        return
    word_times = _build_word_times(segments)
    if not word_times:
        logger.info("Evidence timestamps skipped: segments produced no words (count=%s)", len(segments))
        return
    total = 0
    matched_count = 0
    for result in report.results:
        if result.status != "+":
            continue
        for evidence in result.evidence:
            total += 1
            if evidence.timestamp:
                matched_count += 1
                continue
            matched = _match_timestamp(evidence.quote, word_times)
            if matched:
                evidence.timestamp = matched
                matched_count += 1
    logger.info(
        "Evidence timestamps: segments=%s words=%s evidence=%s matched=%s",
        len(segments),
        len(word_times),
        total,
        matched_count,
    )


def _build_word_times(segments: list[dict[str, Any]]) -> list[tuple[str, float]]:
    out: list[tuple[str, float]] = []
    for seg in segments:
        text = seg.get("text")
        if not isinstance(text, str):
            continue
        try:
            start = float(seg.get("start"))
        except (TypeError, ValueError):
            continue
        for word in _TS_WORD_RE.findall(text.lower()):
            out.append((word, start))
    return out


def _match_timestamp(quote: str, word_times: list[tuple[str, float]]) -> str | None:
    needle = [w.lower() for w in _TS_WORD_RE.findall(quote)][:8]
    if len(needle) < 2:
        return None
    stream = [word for word, _ in word_times]
    best_len = 0
    best_pos = -1
    for i in range(len(stream)):
        run = 0
        while run < len(needle) and i + run < len(stream) and stream[i + run] == needle[run]:
            run += 1
        if run > best_len:
            best_len = run
            best_pos = i
            if best_len == len(needle):
                break
    if best_len >= 2 and best_pos >= 0:
        return _format_mmss(word_times[best_pos][1])
    return None


def _format_mmss(seconds: float) -> str:
    total = max(0, int(round(seconds)))
    return f"{total // 60:02d}:{total % 60:02d}"


def _format_comment(result: IndicatorAnalysis) -> str:
    if result.status != "+":
        return "" if result.status == "-" else result.observable_reason or result.comment

    evidence_lines = []
    for item in result.evidence:
        prefix = f"[{item.timestamp}] " if item.timestamp else ""
        evidence_lines.append(f"{prefix}«{item.quote}»")
    if result.comment:
        evidence_lines.append(result.comment)
    return "\n".join(evidence_lines)


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _detect_notebook_layout(sheet) -> dict[str, int]:
    for row in range(1, sheet.max_row + 1):
        headers = {
            column: _cell_text(sheet.cell(row=row, column=column).value)
            for column in range(1, sheet.max_column + 1)
        }
        status_col = next((column for column, text in headers.items() if text.casefold() == "проявления"), None)
        comment_col = next(
            (column for column, text in headers.items() if text.casefold() in {"комментарии", "комментарий"}),
            None,
        )
        if status_col and comment_col and status_col > 1:
            indicator_col = status_col - 1
            return {
                "competence_col": max(1, indicator_col - 1),
                "indicator_col": indicator_col,
                "status_col": status_col,
                "comment_col": comment_col,
                "level_col": comment_col + 1,
            }

    return {
        "competence_col": 2,
        "indicator_col": 3,
        "status_col": 4,
        "comment_col": 5,
        "level_col": 6,
    }


def _looks_like_level(text: str) -> bool:
    try:
        float(text.replace(",", "."))
    except ValueError:
        return False
    return True


def _level_column_for_competence(indicator_by_row: dict[int, IndicatorRow], competence: str) -> int:
    for indicator in indicator_by_row.values():
        if indicator.competence == competence:
            return indicator.level_col
    return 6


def _looks_like_column_header(text: str) -> bool:
    lowered = text.casefold()
    return lowered in {
        "компетенция",
        "компетенции",
        "поведенческие индикаторы",
        "индикаторы",
        "проявления",
        "комментарии",
        "комментарий",
    }
